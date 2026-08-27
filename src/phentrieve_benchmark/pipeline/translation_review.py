import re
import tempfile
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import TypeVar

from openpyxl import load_workbook  # type: ignore[import-untyped]

from phentrieve_benchmark.artifacts.store import ArtifactStore
from phentrieve_benchmark.models.translation import (
    TranslationManifest,
    TranslationRecord,
)
from phentrieve_benchmark.models.translation_review import (
    ClinicalChange,
    ClinicalChangeCategory,
    TranslationReviewDecision,
    TranslationReviewExport,
    TranslationReviewExportCase,
    TranslationReviewImportEntry,
    TranslationReviewImportManifest,
    TranslationReviewRecord,
)
from phentrieve_benchmark.provenance.canonical import (
    canonical_json_bytes,
    canonical_text_bytes,
)
from phentrieve_benchmark.provenance.digests import sha256_bytes
from phentrieve_benchmark.review.translation_text import unified_text_diff
from phentrieve_benchmark.review.translation_workbook import (
    NMT_HEADER,
    ParsedReviewWorkbook,
    WorkbookCase,
    read_review_workbook,
    write_review_workbook,
)


@dataclass(frozen=True)
class WorkbookValidationIssue:
    sheet: str
    row: int | None
    case_id: str | None
    field: str
    message: str

    def __str__(self) -> str:
        row = str(self.row) if self.row is not None else "-"
        case_id = self.case_id if self.case_id is not None else "-"
        return (
            f"{self.sheet} row {row} case {case_id} field {self.field}: {self.message}"
        )


class WorkbookValidationError(ValueError):
    """All validation failures found in a translation-review workbook."""

    def __init__(self, issues: Sequence[WorkbookValidationIssue]) -> None:
        if not issues:
            raise ValueError("WorkbookValidationError requires at least one issue")
        self.issues = tuple(issues)
        super().__init__("\n".join(str(issue) for issue in self.issues))


@dataclass(frozen=True)
class _PreparedReview:
    source_case_id: str
    proposed_bytes: bytes
    record: TranslationReviewRecord
    review_record_bytes: bytes
    diff_bytes: bytes


_ISO_DATE = re.compile(r"\d{4}-\d{2}-\d{2}")
_EXCEL_UTF16_LIMIT = 32_767
_EnumT = TypeVar(
    "_EnumT", TranslationReviewDecision, ClinicalChange, ClinicalChangeCategory
)
_REVIEW_FIELDS_BY_COLUMN = {
    "A": "source_case_id",
    "B": "source_language",
    "C": "source_text",
    "D": "tllm_text",
    "E": "proposed_text",
    "F": "decision",
    "G": "clinical_change",
    "H": "clinical_change_category",
    "I": "clinical_change_rationale",
    "J": "reviewer_comment",
    "K": "nmt_text",
}
_METADATA_FIELDS_BY_COORDINATE = {
    "B15": "export_sha256",
    "B16": "selection_id",
    "B17": "review_policy_id",
    "B10": "reviewer_id",
    "B11": "reviewer_qualification",
    "B12": "reviewed_languages",
    "B13": "review_date",
}
_EMPTY_METADATA_COORDINATES = {"B10", "B11", "B12", "B13"}


def _records_by_case(
    manifest: TranslationManifest,
    *,
    model: str,
    variant: str,
) -> dict[str, TranslationRecord]:
    if not manifest.records:
        raise ValueError(f"{variant} translation manifest has no cases")
    if any(record.selection_id != manifest.selection_id for record in manifest.records):
        raise ValueError(f"{variant} records do not match the manifest selection ID")
    if any(record.model != model for record in manifest.records):
        raise ValueError(f"{variant} manifest contains the wrong translation model")
    return {record.source_case_id: record for record in manifest.records}


def _filter_language(
    records: dict[str, TranslationRecord],
    *,
    source_language: str | None,
    variant: str,
) -> dict[str, TranslationRecord]:
    """Keep only the cases written in `source_language`.

    A per-language export is its own canonical artifact with its own hash, so
    reviewers of different languages work and import independently.
    """
    if source_language is None:
        return records
    filtered = {
        case_id: record
        for case_id, record in records.items()
        if record.source_language == source_language
    }
    if not filtered:
        raise ValueError(
            f"{variant} manifest has no case in source language {source_language!r}"
        )
    return filtered


def _validate_nmt_manifest(
    tllm_manifest: TranslationManifest,
    nmt_manifest: TranslationManifest,
    *,
    tllm_records: dict[str, TranslationRecord],
    source_language: str | None,
) -> dict[str, TranslationRecord]:
    if nmt_manifest.selection_id != tllm_manifest.selection_id:
        raise ValueError("TLLM and NMT manifests have different selection IDs")
    nmt_records = _filter_language(
        _records_by_case(nmt_manifest, model="general/nmt", variant="NMT"),
        source_language=source_language,
        variant="NMT",
    )
    if nmt_records.keys() != tllm_records.keys():
        raise ValueError("TLLM and NMT manifests have different case sets")
    for case_id, tllm_record in tllm_records.items():
        nmt_record = nmt_records[case_id]
        if nmt_record.source_sha256 != tllm_record.source_sha256:
            raise ValueError("TLLM and NMT manifests have different source hashes")
        if nmt_record.source_language != tllm_record.source_language:
            raise ValueError("TLLM and NMT manifests have different source languages")
    return nmt_records


def export_translation_review(
    *,
    store: ArtifactStore,
    tllm_manifest: TranslationManifest,
    destination: Path,
    review_policy_id: str,
    nmt_manifest: TranslationManifest | None = None,
    source_language: str | None = None,
) -> str:
    """Store a canonical review export and write its Excel review workbook."""
    tllm_records = _filter_language(
        _records_by_case(
            tllm_manifest,
            model="general/translation-llm",
            variant="TLLM",
        ),
        source_language=source_language,
        variant="TLLM",
    )
    nmt_records = (
        _validate_nmt_manifest(
            tllm_manifest,
            nmt_manifest,
            tllm_records=tllm_records,
            source_language=source_language,
        )
        if nmt_manifest is not None
        else None
    )

    export_cases: list[TranslationReviewExportCase] = []
    workbook_cases: list[WorkbookCase] = []
    for tllm_record in sorted(
        tllm_records.values(),
        key=lambda record: (record.source_language, record.source_case_id),
    ):
        source_text = store.read_bytes(tllm_record.source_sha256).decode("utf-8")
        tllm_text = store.read_bytes(tllm_record.translation_sha256).decode("utf-8")
        nmt_record = (
            nmt_records[tllm_record.source_case_id] if nmt_records is not None else None
        )
        nmt_text = None
        if nmt_record is not None:
            store.read_bytes(nmt_record.source_sha256)
            nmt_text = store.read_bytes(nmt_record.translation_sha256).decode("utf-8")
        export_cases.append(
            TranslationReviewExportCase(
                source_case_id=tllm_record.source_case_id,
                source_language=tllm_record.source_language,
                source_text_sha256=tllm_record.source_sha256,
                tllm_text_sha256=tllm_record.translation_sha256,
                nmt_text_sha256=(
                    nmt_record.translation_sha256 if nmt_record is not None else None
                ),
            )
        )
        workbook_cases.append(
            WorkbookCase(
                source_case_id=tllm_record.source_case_id,
                source_language=tllm_record.source_language,
                source_text=source_text,
                tllm_text=tllm_text,
                nmt_text=nmt_text,
            )
        )

    export = TranslationReviewExport(
        selection_id=tllm_manifest.selection_id,
        review_policy_id=review_policy_id,
        nmt_recipe_sha256=(
            nmt_manifest.recipe_sha256 if nmt_manifest is not None else None
        ),
        cases=tuple(export_cases),
    )
    export_sha256 = store.put_bytes(export.canonical_bytes())
    write_review_workbook(destination, export, tuple(workbook_cases))
    return export_sha256


def _issue(
    *,
    sheet: str,
    row: int | None,
    case_id: str | None,
    field: str,
    message: str,
) -> WorkbookValidationIssue:
    return WorkbookValidationIssue(
        sheet=sheet,
        row=row,
        case_id=case_id,
        field=field,
        message=message,
    )


def _has_valid_utf16_length(value: str) -> bool:
    return len(value.encode("utf-16-le")) // 2 <= _EXCEL_UTF16_LIMIT


def _parse_review_date(value: str) -> date | None:
    if _ISO_DATE.fullmatch(value) is None:
        return None
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _cell_issue(
    *,
    workbook_path: Path,
    sheet: str,
    coordinate: str,
    message: str,
) -> WorkbookValidationIssue:
    row = int(re.search(r"\d+", coordinate).group())  # type: ignore[union-attr]
    if sheet == "Anleitung":
        field = _METADATA_FIELDS_BY_COORDINATE.get(coordinate, "workbook")
        case_id = None
    else:
        field = _REVIEW_FIELDS_BY_COLUMN.get(coordinate[0], "workbook")
        case_id = None
        try:
            workbook = load_workbook(
                workbook_path, data_only=False, keep_links=False, read_only=True
            )
            try:
                value = workbook["Review"][f"A{row}"].value
                case_id = value if isinstance(value, str) else None
            finally:
                workbook.close()
        except Exception:
            pass
    return _issue(
        sheet=sheet,
        row=row,
        case_id=case_id,
        field=field,
        message=message,
    )


def _formula_issues(workbook_path: Path) -> tuple[WorkbookValidationIssue, ...]:
    issues: list[WorkbookValidationIssue] = []
    try:
        workbook = load_workbook(
            workbook_path, data_only=False, keep_links=False, read_only=True
        )
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        if cell.data_type == "f":
                            case_id = None
                            if worksheet.title == "Anleitung":
                                field = _METADATA_FIELDS_BY_COORDINATE.get(
                                    cell.coordinate, "workbook"
                                )
                            elif worksheet.title == "Review":
                                field = _REVIEW_FIELDS_BY_COLUMN.get(
                                    cell.column_letter, "workbook"
                                )
                                value = worksheet[f"A{cell.row}"].value
                                case_id = value if isinstance(value, str) else None
                            else:
                                field = "workbook"
                            issues.append(
                                _issue(
                                    sheet=worksheet.title,
                                    row=cell.row,
                                    case_id=case_id,
                                    field=field,
                                    message="formula cells are forbidden",
                                )
                            )
        finally:
            workbook.close()
    except Exception:
        return ()
    return tuple(issues)


def _invalid_consumed_string(value: object, *, allow_empty: bool) -> bool:
    return not isinstance(value, str) and not (allow_empty and value is None)


def _consumed_non_string_issues(
    workbook_path: Path,
) -> tuple[WorkbookValidationIssue, ...]:
    issues: list[WorkbookValidationIssue] = []
    try:
        workbook = load_workbook(
            workbook_path, data_only=False, keep_links=False, read_only=True
        )
        try:
            if "Anleitung" in workbook.sheetnames:
                instructions = workbook["Anleitung"]
                for coordinate, field in _METADATA_FIELDS_BY_COORDINATE.items():
                    value = instructions[coordinate].value
                    if _invalid_consumed_string(
                        value,
                        allow_empty=coordinate in _EMPTY_METADATA_COORDINATES,
                    ):
                        issues.append(
                            _issue(
                                sheet="Anleitung",
                                row=instructions[coordinate].row,
                                case_id=None,
                                field=field,
                                message="must contain a string",
                            )
                        )
            if "Review" in workbook.sheetnames:
                review = workbook["Review"]
                last_column = 11 if review["K1"].value == NMT_HEADER else 10
                for row_number in range(2, review.max_row + 1):
                    case_value = review[f"A{row_number}"].value
                    case_id = case_value if isinstance(case_value, str) else None
                    for column_number in range(1, last_column + 1):
                        cell = review.cell(row_number, column_number)
                        if _invalid_consumed_string(
                            cell.value,
                            allow_empty=6 <= column_number <= 10,
                        ):
                            issues.append(
                                _issue(
                                    sheet="Review",
                                    row=row_number,
                                    case_id=case_id,
                                    field=_REVIEW_FIELDS_BY_COLUMN[cell.column_letter],
                                    message="must contain a string",
                                )
                            )
        finally:
            workbook.close()
    except Exception:
        return ()
    return tuple(issues)


def _non_string_issue(
    *, workbook_path: Path, coordinate: str
) -> WorkbookValidationIssue:
    sheet = "Review"
    if coordinate in _METADATA_FIELDS_BY_COORDINATE:
        try:
            workbook = load_workbook(
                workbook_path, data_only=False, keep_links=False, read_only=True
            )
            try:
                metadata_value = workbook["Anleitung"][coordinate].value
                metadata_is_invalid = not isinstance(metadata_value, str) and not (
                    metadata_value is None and coordinate in _EMPTY_METADATA_COORDINATES
                )
                if metadata_is_invalid:
                    sheet = "Anleitung"
            finally:
                workbook.close()
        except Exception:
            pass
    return _cell_issue(
        workbook_path=workbook_path,
        sheet=sheet,
        coordinate=coordinate,
        message="must contain a string",
    )


def _parser_error(error: Exception, *, workbook_path: Path) -> WorkbookValidationError:
    message = str(error)
    if "exactly Anleitung and Review" in message:
        message = f"extra or missing sheet; {message}"
    cell_error = re.fullmatch(r"([A-K])(\d+) must contain a string", message)
    if cell_error is not None:
        column, row_text = cell_error.groups()
        coordinate = f"{column}{row_text}"
        return WorkbookValidationError(
            (
                _non_string_issue(
                    workbook_path=workbook_path,
                    coordinate=coordinate,
                ),
            )
        )
    return WorkbookValidationError(
        (
            _issue(
                sheet="workbook",
                row=None,
                case_id=None,
                field="workbook",
                message=message,
            ),
        )
    )


def _read_without_parser_cells(
    workbook_path: Path,
) -> tuple[ParsedReviewWorkbook | None, tuple[WorkbookValidationIssue, ...]]:
    workbook = load_workbook(workbook_path, data_only=False, keep_links=False)
    temporary_path: Path | None = None
    try:
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        cell.value = (
                            f"formula removed from {worksheet.title}!{cell.coordinate}"
                        )
        if "Anleitung" in workbook.sheetnames:
            instructions = workbook["Anleitung"]
            for coordinate in _METADATA_FIELDS_BY_COORDINATE:
                value = instructions[coordinate].value
                if _invalid_consumed_string(
                    value,
                    allow_empty=coordinate in _EMPTY_METADATA_COORDINATES,
                ):
                    instructions[coordinate] = (
                        ""
                        if coordinate in _EMPTY_METADATA_COORDINATES
                        else "invalid non-string cell"
                    )
        if "Review" in workbook.sheetnames:
            review = workbook["Review"]
            last_column = 11 if review["K1"].value == NMT_HEADER else 10
            for row_number in range(2, review.max_row + 1):
                for column_number in range(1, last_column + 1):
                    cell = review.cell(row_number, column_number)
                    if _invalid_consumed_string(
                        cell.value,
                        allow_empty=6 <= column_number <= 10,
                    ):
                        cell.value = (
                            ""
                            if 6 <= column_number <= 10
                            else "invalid non-string cell"
                        )
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temporary:
            temporary_path = Path(temporary.name)
        workbook.save(temporary_path)
        try:
            return read_review_workbook(temporary_path), ()
        except Exception as error:
            return None, _parser_error(error, workbook_path=temporary_path).issues
    finally:
        workbook.close()
        if temporary_path is not None:
            temporary_path.unlink(missing_ok=True)


def _read_workbook_for_import(
    workbook_path: Path,
) -> tuple[ParsedReviewWorkbook | None, list[WorkbookValidationIssue]]:
    try:
        return read_review_workbook(workbook_path), []
    except Exception as error:
        formula_issues = _formula_issues(workbook_path)
        non_string_issues = _consumed_non_string_issues(workbook_path)
        parser_cell_issues = [*formula_issues, *non_string_issues]
        if not parser_cell_issues:
            return None, list(_parser_error(error, workbook_path=workbook_path).issues)
        try:
            workbook, parser_issues = _read_without_parser_cells(workbook_path)
        except Exception as sanitized_error:
            parser_issues = _parser_error(
                sanitized_error, workbook_path=workbook_path
            ).issues
            workbook = None
        return workbook, [*parser_cell_issues, *parser_issues]


def _load_export(
    *,
    store: ArtifactStore,
    export_sha256: str,
    issues: list[WorkbookValidationIssue],
) -> TranslationReviewExport | None:
    try:
        export_bytes = store.read_bytes(export_sha256)
        export = TranslationReviewExport.model_validate_json(export_bytes, strict=True)
        if export.canonical_bytes() != export_bytes:
            raise ValueError("export manifest bytes are not canonical")
    except Exception as error:
        issues.append(
            _issue(
                sheet="Anleitung",
                row=3,
                case_id=None,
                field="export_sha256",
                message=f"does not resolve to a valid canonical export: {error}",
            )
        )
        return None
    return export


def _validate_metadata(
    *,
    workbook: object,
    export: TranslationReviewExport | None,
    issues: list[WorkbookValidationIssue],
) -> date | None:
    metadata = (
        ("selection_id", 4, workbook.selection_id),  # type: ignore[attr-defined]
        ("review_policy_id", 5, workbook.review_policy_id),  # type: ignore[attr-defined]
        ("reviewer_id", 7, workbook.reviewer_id),  # type: ignore[attr-defined]
        (
            "reviewer_qualification",
            8,
            workbook.reviewer_qualification,  # type: ignore[attr-defined]
        ),
        ("reviewed_languages", 9, workbook.reviewed_languages),  # type: ignore[attr-defined]
        ("review_date", 10, workbook.review_date),  # type: ignore[attr-defined]
    )
    for field, row, value in metadata:
        if not value:
            issues.append(
                _issue(
                    sheet="Anleitung",
                    row=row,
                    case_id=None,
                    field=field,
                    message="is required",
                )
            )
        if not _has_valid_utf16_length(value):
            issues.append(
                _issue(
                    sheet="Anleitung",
                    row=row,
                    case_id=None,
                    field=field,
                    message="exceeds Excel's 32,767 UTF-16 code-unit limit",
                )
            )

    if export is not None:
        if workbook.selection_id != export.selection_id:  # type: ignore[attr-defined]
            issues.append(
                _issue(
                    sheet="Anleitung",
                    row=4,
                    case_id=None,
                    field="selection_id",
                    message="does not match the authoritative export",
                )
            )
        if workbook.review_policy_id != export.review_policy_id:  # type: ignore[attr-defined]
            issues.append(
                _issue(
                    sheet="Anleitung",
                    row=5,
                    case_id=None,
                    field="review_policy_id",
                    message="does not match the authoritative export",
                )
            )

    review_date = _parse_review_date(workbook.review_date)  # type: ignore[attr-defined]
    if review_date is None and workbook.review_date:  # type: ignore[attr-defined]
        issues.append(
            _issue(
                sheet="Anleitung",
                row=10,
                case_id=None,
                field="review_date",
                message="must be a valid date in YYYY-MM-DD form",
            )
        )
    return review_date


def _read_export_text(
    *,
    store: ArtifactStore,
    digest: str,
    row: int,
    case_id: str,
    field: str,
    issues: list[WorkbookValidationIssue],
) -> str | None:
    try:
        artifact_bytes = store.read_bytes(digest)
        text = artifact_bytes.decode("utf-8")
    except Exception as error:
        issues.append(
            _issue(
                sheet="Review",
                row=row,
                case_id=case_id,
                field=field,
                message=f"authoritative artifact cannot be read: {error}",
            )
        )
        return None
    if canonical_text_bytes(text) != artifact_bytes:
        issues.append(
            _issue(
                sheet="Review",
                row=row,
                case_id=case_id,
                field=field,
                message="authoritative artifact is not canonical text bytes",
            )
        )
        return None
    return text


def _validate_enum(
    value: str,
    enum_type: type[_EnumT],
    *,
    row: int,
    case_id: str,
    field: str,
    required: bool,
    issues: list[WorkbookValidationIssue],
) -> _EnumT | None:
    if not value:
        if required:
            issues.append(
                _issue(
                    sheet="Review",
                    row=row,
                    case_id=case_id,
                    field=field,
                    message="is required",
                )
            )
        return None
    try:
        return enum_type(value)
    except ValueError:
        issues.append(
            _issue(
                sheet="Review",
                row=row,
                case_id=case_id,
                field=field,
                message="is not an allowed value",
            )
        )
        return None


def _decision_is_allowed(
    *,
    decision: TranslationReviewDecision,
    changed: bool,
    clinical_change: ClinicalChange,
    category: ClinicalChangeCategory | None,
    rationale: str | None,
) -> bool:
    has_details = category is not None and rationale is not None
    has_no_details = category is None and rationale is None
    if decision is TranslationReviewDecision.ACCEPTED_UNCHANGED:
        return not changed and clinical_change is ClinicalChange.NONE and has_no_details
    if decision is TranslationReviewDecision.ACCEPTED_CORRECTED:
        return changed and (
            (clinical_change is ClinicalChange.NONE and has_no_details)
            or (clinical_change is ClinicalChange.PRESENT and has_details)
        )
    return clinical_change is ClinicalChange.PRESENT and has_details


def _validate_rows(
    *,
    store: ArtifactStore,
    workbook: object,
    export: TranslationReviewExport | None,
    review_date: date | None,
    issues: list[WorkbookValidationIssue],
) -> list[_PreparedReview]:
    prepared: list[_PreparedReview] = []
    rows = workbook.rows  # type: ignore[attr-defined]
    seen: dict[str, int] = {}
    export_cases = (
        {case.source_case_id: case for case in export.cases}
        if export is not None
        else {}
    )
    includes_nmt = export is not None and export.nmt_recipe_sha256 is not None
    workbook_includes_nmt = any(row.nmt_text is not None for row in rows)
    if export is not None and workbook_includes_nmt != includes_nmt:
        issues.append(
            _issue(
                sheet="Review",
                row=1,
                case_id=None,
                field="nmt_text",
                message="NMT column presence does not match the export",
            )
        )

    for row_number, row in enumerate(rows, start=2):
        case_id = row.source_case_id
        issue_count = len(issues)
        if case_id in seen:
            issues.append(
                _issue(
                    sheet="Review",
                    row=row_number,
                    case_id=case_id,
                    field="source_case_id",
                    message=f"duplicates row {seen[case_id]}",
                )
            )
        else:
            seen[case_id] = row_number
        export_case = export_cases.get(case_id)
        if export is not None and export_case is None:
            issues.append(
                _issue(
                    sheet="Review",
                    row=row_number,
                    case_id=case_id,
                    field="source_case_id",
                    message="is not part of the authoritative export",
                )
            )

        row_text: list[tuple[str, str]] = [
            ("source_case_id", row.source_case_id),
            ("source_language", row.source_language),
            ("source_text", row.source_text),
            ("tllm_text", row.tllm_text),
            ("proposed_text", row.proposed_text),
            ("decision", row.decision),
            ("clinical_change", row.clinical_change),
            ("clinical_change_category", row.clinical_change_category),
            ("clinical_change_rationale", row.clinical_change_rationale),
            ("reviewer_comment", row.reviewer_comment),
        ]
        if row.nmt_text is not None:
            row_text.append(("nmt_text", row.nmt_text))
        for field, value in row_text:
            if not _has_valid_utf16_length(value):
                issues.append(
                    _issue(
                        sheet="Review",
                        row=row_number,
                        case_id=case_id,
                        field=field,
                        message="exceeds Excel's 32,767 UTF-16 code-unit limit",
                    )
                )

        if not row.proposed_text:
            issues.append(
                _issue(
                    sheet="Review",
                    row=row_number,
                    case_id=case_id,
                    field="proposed_text",
                    message="is required",
                )
            )

        decision = _validate_enum(
            row.decision,
            TranslationReviewDecision,
            row=row_number,
            case_id=case_id,
            field="decision",
            required=True,
            issues=issues,
        )
        clinical_change = _validate_enum(
            row.clinical_change,
            ClinicalChange,
            row=row_number,
            case_id=case_id,
            field="clinical_change",
            required=True,
            issues=issues,
        )
        category = _validate_enum(
            row.clinical_change_category,
            ClinicalChangeCategory,
            row=row_number,
            case_id=case_id,
            field="clinical_change_category",
            required=False,
            issues=issues,
        )
        rationale = row.clinical_change_rationale or None

        if export_case is not None:
            if row.source_language != export_case.source_language:
                issues.append(
                    _issue(
                        sheet="Review",
                        row=row_number,
                        case_id=case_id,
                        field="source_language",
                        message="does not match the authoritative export",
                    )
                )
            expected_source = _read_export_text(
                store=store,
                digest=export_case.source_text_sha256,
                row=row_number,
                case_id=case_id,
                field="source_text",
                issues=issues,
            )
            expected_tllm = _read_export_text(
                store=store,
                digest=export_case.tllm_text_sha256,
                row=row_number,
                case_id=case_id,
                field="tllm_text",
                issues=issues,
            )
            expected_nmt = (
                _read_export_text(
                    store=store,
                    digest=export_case.nmt_text_sha256,
                    row=row_number,
                    case_id=case_id,
                    field="nmt_text",
                    issues=issues,
                )
                if export_case.nmt_text_sha256 is not None
                else None
            )
            for field, actual, expected in (
                ("source_text", row.source_text, expected_source),
                ("tllm_text", row.tllm_text, expected_tllm),
                ("nmt_text", row.nmt_text, expected_nmt),
            ):
                if expected is not None and (
                    actual is None
                    or canonical_text_bytes(actual) != canonical_text_bytes(expected)
                ):
                    issues.append(
                        _issue(
                            sheet="Review",
                            row=row_number,
                            case_id=case_id,
                            field=field,
                            message="does not match the authoritative artifact",
                        )
                    )

            changed = (
                canonical_text_bytes(row.proposed_text)
                != canonical_text_bytes(expected_tllm)
                if expected_tllm is not None
                else None
            )
            if (
                decision is not None
                and clinical_change is not None
                and changed is not None
                and not _decision_is_allowed(
                    decision=decision,
                    changed=changed,
                    clinical_change=clinical_change,
                    category=category,
                    rationale=rationale,
                )
            ):
                issues.append(
                    _issue(
                        sheet="Review",
                        row=row_number,
                        case_id=case_id,
                        field="decision",
                        message="field combination violates the decision table",
                    )
                )

            if (
                issue_count == 0
                and len(issues) == issue_count
                and review_date is not None
                and decision is not None
                and clinical_change is not None
                and export is not None
            ):
                proposed_bytes = canonical_text_bytes(row.proposed_text)
                record = TranslationReviewRecord(
                    export_sha256=workbook.export_sha256,  # type: ignore[attr-defined]
                    source_case_id=case_id,
                    source_language=export_case.source_language,
                    target_language="de",
                    source_text_sha256=export_case.source_text_sha256,
                    tllm_text_sha256=export_case.tllm_text_sha256,
                    proposed_text_sha256=sha256_bytes(proposed_bytes),
                    reviewer_id=workbook.reviewer_id,  # type: ignore[attr-defined]
                    reviewer_qualification=workbook.reviewer_qualification,  # type: ignore[attr-defined]
                    reviewed_languages=workbook.reviewed_languages,  # type: ignore[attr-defined]
                    review_date=review_date,
                    review_policy_id=export.review_policy_id,
                    decision=decision,
                    clinical_change=clinical_change,
                    clinical_change_category=category,
                    clinical_change_rationale=rationale,
                    reviewer_comment=row.reviewer_comment or None,
                )
                diff = unified_text_diff(row.tllm_text, row.proposed_text)
                prepared.append(
                    _PreparedReview(
                        source_case_id=case_id,
                        proposed_bytes=proposed_bytes,
                        record=record,
                        review_record_bytes=canonical_json_bytes(
                            record.review_record().model_dump(mode="json")
                        ),
                        diff_bytes=diff.canonical_bytes(),
                    )
                )

    if export is not None:
        for missing_case_id in export_cases.keys() - seen.keys():
            issues.append(
                _issue(
                    sheet="Review",
                    row=None,
                    case_id=missing_case_id,
                    field="source_case_id",
                    message="export case is missing from the workbook",
                )
            )
    return prepared


def import_translation_review(*, store: ArtifactStore, workbook_path: Path) -> str:
    """Validate fully, then publish a completed workbook manifest last."""
    workbook, issues = _read_workbook_for_import(workbook_path)
    if workbook is None:
        raise WorkbookValidationError(issues)
    export = _load_export(
        store=store,
        export_sha256=workbook.export_sha256,
        issues=issues,
    )
    review_date = _validate_metadata(
        workbook=workbook,
        export=export,
        issues=issues,
    )
    prepared = _validate_rows(
        store=store,
        workbook=workbook,
        export=export,
        review_date=review_date,
        issues=issues,
    )
    if issues:
        raise WorkbookValidationError(issues)

    entries: list[TranslationReviewImportEntry] = []
    for item in prepared:
        proposed_text_sha256 = store.put_bytes(item.proposed_bytes)
        record_sha256 = store.put_bytes(item.record.canonical_bytes())
        review_record_sha256 = store.put_bytes(item.review_record_bytes)
        diff_sha256 = store.put_bytes(item.diff_bytes)
        entries.append(
            TranslationReviewImportEntry(
                source_case_id=item.source_case_id,
                proposed_text_sha256=proposed_text_sha256,
                record_sha256=record_sha256,
                review_record_sha256=review_record_sha256,
                diff_sha256=diff_sha256,
            )
        )

    manifest = TranslationReviewImportManifest(
        export_sha256=workbook.export_sha256,
        entries=tuple(entries),
    )
    return store.put_bytes(manifest.canonical_bytes())
