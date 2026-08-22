import os
import tempfile
from dataclasses import dataclass
from pathlib import Path

import openpyxl  # type: ignore[import-untyped]
from openpyxl.styles import Alignment, Font, Protection  # type: ignore[import-untyped]
from openpyxl.worksheet.datavalidation import (  # type: ignore[import-untyped]
    DataValidation,
)

from phentrieve_benchmark.models.translation_review import TranslationReviewExport

REVIEW_HEADERS = (
    "Fall-ID",
    "Quellsprache",
    "Originaltext",
    "TLLM-Ausgangsfassung",
    "Korrigierte Endfassung",
    "Entscheidung",
    "Klinisch relevante Änderung",
    "Hauptkategorie",
    "Änderungsbegründung",
    "Reviewer-Kommentar",
)
NMT_HEADER = "NMT-Vergleich"
DECISION_VALUES = (
    "unverändert akzeptiert",
    "korrigiert akzeptiert",
    "Rückfrage",
    "abgelehnt",
)
CLINICAL_CHANGE_VALUES = ("keine", "vorhanden")
CATEGORY_VALUES = (
    "Auslassung",
    "Hinzufügung",
    "Negation oder Aussagesicherheit",
    "Zahl oder Einheit",
    "Anatomie oder Lateralität",
    "Terminologie",
    "Quellproblem",
)

_METADATA_CELLS = {
    "export_sha256": "B15",
    "selection_id": "B16",
    "review_policy_id": "B17",
    "reviewer_id": "B10",
    "reviewer_qualification": "B11",
    "reviewed_languages": "B12",
    "review_date": "B13",
}
_EDITABLE_REVIEW_COLUMNS = range(5, 11)


@dataclass(frozen=True)
class WorkbookCase:
    source_case_id: str
    source_language: str
    source_text: str
    tllm_text: str
    nmt_text: str | None = None


@dataclass(frozen=True)
class ParsedReviewRow:
    source_case_id: str
    source_language: str
    source_text: str
    tllm_text: str
    proposed_text: str
    decision: str
    clinical_change: str
    clinical_change_category: str
    clinical_change_rationale: str
    reviewer_comment: str
    nmt_text: str | None


@dataclass(frozen=True)
class ParsedReviewWorkbook:
    export_sha256: str
    selection_id: str
    review_policy_id: str
    reviewer_id: str
    reviewer_qualification: str
    reviewed_languages: str
    review_date: str
    rows: tuple[ParsedReviewRow, ...]


def _utf16_code_units(value: str) -> int:
    return len(value.encode("utf-16-le")) // 2


def _validate_exportable_text(value: str, *, field: str) -> None:
    if not isinstance(value, str):
        raise ValueError(f"{field} must be a string")
    if _utf16_code_units(value) > 32_767:
        raise ValueError(f"{field} exceeds Excel's 32,767 UTF-16 code-unit limit")


def _write_text(cell: object, value: str) -> None:
    _validate_exportable_text(value, field="workbook cell")
    cell.value = value  # type: ignore[attr-defined]
    cell.data_type = "s"  # type: ignore[attr-defined]


def _add_list_validation(
    worksheet: object, *, values: tuple[str, ...], cell_range: str
) -> None:
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=True,
    )
    worksheet.add_data_validation(validation)  # type: ignore[attr-defined]
    validation.add(cell_range)


def _format_review_sheet(worksheet: object, *, includes_nmt: bool, rows: int) -> None:
    worksheet.freeze_panes = "A2"  # type: ignore[attr-defined]
    last_column = "K" if includes_nmt else "J"
    worksheet.auto_filter.ref = f"A1:{last_column}{rows + 1}"  # type: ignore[attr-defined]
    worksheet.protection.sheet = True  # type: ignore[attr-defined]
    worksheet.protection.autoFilter = False  # type: ignore[attr-defined]
    worksheet.protection.selectLockedCells = False  # type: ignore[attr-defined]
    worksheet.protection.selectUnlockedCells = False  # type: ignore[attr-defined]

    widths = (18, 14, 48, 48, 48, 28, 28, 34, 42, 36, 48)
    for index, width in enumerate(widths[: 11 if includes_nmt else 10], start=1):
        worksheet.column_dimensions[  # type: ignore[attr-defined]
            openpyxl.utils.get_column_letter(index)
        ].width = width

    for row_index in range(1, rows + 2):
        worksheet.row_dimensions[row_index].height = 75 if row_index > 1 else 30  # type: ignore[attr-defined]
        for column_index in range(1, 12 if includes_nmt else 11):
            cell = worksheet.cell(row_index, column_index)  # type: ignore[attr-defined]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_index == 1:
                cell.font = Font(bold=True)
            elif column_index in _EDITABLE_REVIEW_COLUMNS:
                cell.protection = Protection(locked=False)

    if rows:
        _add_list_validation(
            worksheet,
            values=DECISION_VALUES,
            cell_range=f"F2:F{rows + 1}",
        )
        _add_list_validation(
            worksheet,
            values=CLINICAL_CHANGE_VALUES,
            cell_range=f"G2:G{rows + 1}",
        )
        _add_list_validation(
            worksheet,
            values=CATEGORY_VALUES,
            cell_range=f"H2:H{rows + 1}",
        )


def _format_instructions_sheet(worksheet: object) -> None:
    worksheet.column_dimensions["A"].width = 34  # type: ignore[attr-defined]
    worksheet.column_dimensions["B"].width = 88  # type: ignore[attr-defined]
    worksheet.protection.sheet = True  # type: ignore[attr-defined]
    worksheet.protection.selectLockedCells = False  # type: ignore[attr-defined]
    worksheet.protection.selectUnlockedCells = False  # type: ignore[attr-defined]
    for row in worksheet.iter_rows():  # type: ignore[attr-defined]
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell_range in ("A1:B1", "A3:B3", "A4:B4", "A5:B5", "A6:B6", "A7:B7", "A8:B8"):
        worksheet.merge_cells(cell_range)  # type: ignore[attr-defined]
    for coordinate in ("B10", "B11", "B12", "B13"):
        worksheet[coordinate].protection = Protection(locked=False)  # type: ignore[index]
    worksheet["B13"].number_format = "@"  # type: ignore[index]
    date_validation = DataValidation(
        type="custom",
        formula1='=AND(ISTEXT(B13),LEN(B13)=10,MID(B13,5,1)="-",MID(B13,8,1)="-")',
        allow_blank=True,
    )
    worksheet.add_data_validation(date_validation)  # type: ignore[attr-defined]
    date_validation.add("B13")


def _validate_cases(cases: tuple[WorkbookCase, ...], *, includes_nmt: bool) -> None:
    if not cases:
        raise ValueError("review workbook requires at least one case")
    for case in cases:
        for field, value in (
            ("source case ID", case.source_case_id),
            ("source language", case.source_language),
            ("source text", case.source_text),
            ("TLLM text", case.tllm_text),
        ):
            _validate_exportable_text(value, field=field)
        if includes_nmt:
            if case.nmt_text is None:
                raise ValueError("NMT export requires NMT text for every case")
            _validate_exportable_text(case.nmt_text, field="NMT text")
        elif case.nmt_text is not None:
            raise ValueError("NMT text requires an NMT export")


def write_review_workbook(
    destination: Path,
    export: TranslationReviewExport,
    cases: tuple[WorkbookCase, ...],
) -> None:
    if destination.suffix.casefold() != ".xlsx":
        raise ValueError("review workbook destination must end in .xlsx")
    includes_nmt = export.nmt_recipe_sha256 is not None
    _validate_cases(cases, includes_nmt=includes_nmt)
    for field, value in (
        ("export ID", export.sha256()),
        ("selection ID", export.selection_id),
        ("review policy ID", export.review_policy_id),
    ):
        _validate_exportable_text(value, field=field)

    workbook = openpyxl.Workbook()
    instructions = workbook.active
    instructions.title = "Anleitung"
    review = workbook.create_sheet("Review")

    _write_text(instructions["A1"], "E3C-Übersetzungsreview")
    instructions["A1"].font = Font(bold=True)
    _write_text(instructions["A3"], "Arbeitsablauf")
    instructions["A3"].font = Font(bold=True)
    _write_text(
        instructions["A4"],
        "1. Metadaten ausfüllen, dann Originaltext und TLLM-Ausgangsfassung jeder "
        "Zeile vergleichen und die Korrigierte Endfassung als vollständigen "
        "deutschen Endtext bearbeiten (kein Patch, keine isolierte Ersatzphrase). "
        "Für längere, durch die feste Zeilenhöhe verdeckte Texte die "
        "Bearbeitungsleiste oder den Zelleditor verwenden.",
    )
    _write_text(
        instructions["A5"],
        "2. Entscheidung wählen: unverändert akzeptiert, korrigiert akzeptiert, "
        "Rückfrage, wenn sich Quelle oder korrekte Wiedergabe nicht auflösen "
        "lassen, oder abgelehnt, wenn der Fall durch Review nicht geeignet "
        "gemacht werden kann.",
    )
    _write_text(
        instructions["A6"],
        "3. Bei vorhanden Hauptkategorie und Änderungsbegründung ausfüllen; bei "
        "keine beide Felder leer lassen.",
    )
    _write_text(
        instructions["A7"],
        "4. Unverändert akzeptiert ist nur bei identischer TLLM-Endfassung mit "
        "keine erlaubt. Korrigiert akzeptiert erfordert eine geänderte Endfassung; "
        "Rückfrage und abgelehnt erfordern vorhanden.",
    )
    _write_text(
        instructions["A8"],
        "5. Ohne Formeln, zusätzliche Blätter oder Dateitypwechsel als .xlsx "
        "speichern.",
    )
    for row in range(4, 9):
        instructions.row_dimensions[row].height = 45
    _write_text(instructions["A9"], "Reviewer-Angaben")
    instructions["A9"].font = Font(bold=True)
    _write_text(instructions["A10"], "Reviewer-ID oder Name")
    _write_text(instructions["A11"], "Medizinische Qualifikation")
    _write_text(instructions["A12"], "Überprüfte Sprachen")
    _write_text(instructions["A13"], "Review-Datum (YYYY-MM-DD)")
    _write_text(instructions["A14"], "Technische Angaben")
    instructions["A14"].font = Font(bold=True)
    _write_text(instructions["A15"], "Export-ID")
    _write_text(instructions["B15"], export.sha256())
    _write_text(instructions["A16"], "Auswahl-ID")
    _write_text(instructions["B16"], export.selection_id)
    _write_text(instructions["A17"], "Review-Policy-ID")
    _write_text(instructions["B17"], export.review_policy_id)

    headers = (*REVIEW_HEADERS, NMT_HEADER) if includes_nmt else REVIEW_HEADERS
    for column_index, header in enumerate(headers, start=1):
        _write_text(review.cell(1, column_index), header)
    for row_index, case in enumerate(cases, start=2):
        row_values: tuple[str, ...] = (
            case.source_case_id,
            case.source_language,
            case.source_text,
            case.tllm_text,
            case.tllm_text,
            "",
            "",
            "",
            "",
            "",
        )
        if includes_nmt:
            row_values += (case.nmt_text or "",)
        for column_index, value in enumerate(row_values, start=1):
            _write_text(review.cell(row_index, column_index), value)

    _format_instructions_sheet(instructions)
    _format_review_sheet(review, includes_nmt=includes_nmt, rows=len(cases))

    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            suffix=".xlsx", dir=destination.parent, delete=False
        ) as temporary_file:
            temporary_path = Path(temporary_file.name)
        workbook.save(temporary_path)
        os.replace(temporary_path, destination)
    finally:
        workbook.close()
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()


def _read_string(
    worksheet: object,
    coordinate: str,
    *,
    empty_is_string: bool = False,
) -> str:
    value = worksheet[coordinate].value  # type: ignore[index]
    if value is None and empty_is_string:
        return ""
    if not isinstance(value, str):
        raise ValueError(f"{coordinate} must contain a string")
    return value


def _reject_formulas(workbook: object) -> None:
    for worksheet in workbook.worksheets:  # type: ignore[attr-defined]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise ValueError("formula cells are forbidden")


def read_review_workbook(source: Path) -> ParsedReviewWorkbook:
    if source.suffix.casefold() != ".xlsx":
        raise ValueError("review workbook source must end in .xlsx")
    workbook = openpyxl.load_workbook(source, data_only=False, keep_links=False)
    try:
        if workbook.sheetnames != ["Anleitung", "Review"]:
            raise ValueError(
                "review workbook must contain exactly Anleitung and Review"
            )
        _reject_formulas(workbook)
        instructions = workbook["Anleitung"]
        review = workbook["Review"]
        actual_headers = tuple(review.cell(1, column).value for column in range(1, 11))
        if actual_headers != REVIEW_HEADERS:
            raise ValueError("review workbook headers do not match the profile")
        optional_header = review.cell(1, 11).value
        if optional_header not in {None, NMT_HEADER}:
            raise ValueError("review workbook headers do not match the profile")
        includes_nmt = optional_header == NMT_HEADER

        metadata = {
            name: _read_string(
                instructions,
                coordinate,
                empty_is_string=name
                in {
                    "reviewer_id",
                    "reviewer_qualification",
                    "reviewed_languages",
                    "review_date",
                },
            )
            for name, coordinate in _METADATA_CELLS.items()
        }
        rows: list[ParsedReviewRow] = []
        expected_columns = 11 if includes_nmt else 10
        data_rows = (
            row_index
            for row_index in range(2, review.max_row + 1)
            if any(
                review.cell(row_index, column_index).value is not None
                for column_index in range(1, expected_columns + 1)
            )
        )
        for row_index in data_rows:
            values = [
                _read_string(
                    review,
                    f"{openpyxl.utils.get_column_letter(column_index)}{row_index}",
                    empty_is_string=column_index >= 6,
                )
                for column_index in range(1, 11)
            ]
            nmt_text = _read_string(review, f"K{row_index}") if includes_nmt else None
            rows.append(
                ParsedReviewRow(
                    source_case_id=values[0],
                    source_language=values[1],
                    source_text=values[2],
                    tllm_text=values[3],
                    proposed_text=values[4],
                    decision=values[5],
                    clinical_change=values[6],
                    clinical_change_category=values[7],
                    clinical_change_rationale=values[8],
                    reviewer_comment=values[9],
                    nmt_text=nmt_text,
                )
            )
        return ParsedReviewWorkbook(rows=tuple(rows), **metadata)
    finally:
        workbook.close()
