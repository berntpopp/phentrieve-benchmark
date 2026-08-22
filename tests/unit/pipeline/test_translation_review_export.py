from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import load_workbook

from phentrieve_benchmark.artifacts.store import ArtifactStore
from phentrieve_benchmark.models.translation import (
    TranslationManifest,
    TranslationRecord,
    TranslationStatus,
)
from phentrieve_benchmark.models.translation_review import TranslationReviewExport
from phentrieve_benchmark.pipeline.translation_review import (
    export_translation_review,
)

_CASES = (
    ("FR2", "fr", "Symptôme français.", "TLLM français.", "NMT français."),
    ("EN1", "en", "English symptom.", "TLLM English.", "NMT English."),
    ("ES3", "es", "Síntoma español.", "TLLM español.", "NMT español."),
)


def _manifest(
    store: ArtifactStore,
    *,
    model: str,
    selection_id: str = "selection-1",
    cases: tuple[tuple[str, str, str, str, str], ...] = _CASES,
) -> TranslationManifest:
    records = tuple(
        TranslationRecord(
            translation_id=f"{model}-{case_id}",
            selection_id=selection_id,
            source_case_id=case_id,
            source_language=language,  # type: ignore[arg-type]
            target_language="de",
            source_sha256=store.put_bytes(source.encode("utf-8")),
            translation_sha256=store.put_bytes(
                (tllm if model == "general/translation-llm" else nmt).encode("utf-8")
            ),
            provider="google-cloud-translation",
            api_version="v3",
            model=model,  # type: ignore[arg-type]
            project_id="phentrieve",
            location="global",
            created_at=datetime(2026, 8, 22, tzinfo=UTC),
            input_codepoints=len(source),
            output_codepoints=len(tllm if model == "general/translation-llm" else nmt),
            price_per_million_input_characters=Decimal("10"),
            estimated_max_cost=Decimal("0"),
            status=TranslationStatus.TRANSLATED,
            checks=(),
        )
        for case_id, language, source, tllm, nmt in cases
    )
    manifest = TranslationManifest(
        selection_id=selection_id,
        selection_sha256="a" * 64,
        recipe_sha256=("b" if model == "general/translation-llm" else "c") * 64,
        records=records,
    )
    store.put_bytes(manifest.canonical_bytes())
    return manifest


def _manifests(store: ArtifactStore) -> tuple[TranslationManifest, TranslationManifest]:
    return (
        _manifest(store, model="general/translation-llm"),
        _manifest(store, model="general/nmt"),
    )


def test_export_binds_store_text_and_full_tllm_hashes_in_canonical_order(
    tmp_path: Path,
) -> None:
    store = ArtifactStore(tmp_path / "objects")
    tllm, _ = _manifests(store)
    output = tmp_path / "review.xlsx"

    export_sha256 = export_translation_review(
        store=store,
        tllm_manifest=tllm,
        destination=output,
        review_policy_id="medical-review-v1",
    )

    export = TranslationReviewExport.model_validate_json(
        store.read_bytes(export_sha256), strict=True
    )
    records = {record.source_case_id: record for record in tllm.records}
    assert [case.source_case_id for case in export.cases] == ["EN1", "ES3", "FR2"]
    assert all(case.nmt_text_sha256 is None for case in export.cases)
    assert export.nmt_recipe_sha256 is None
    assert [case.source_text_sha256 for case in export.cases] == [
        records["EN1"].source_sha256,
        records["ES3"].source_sha256,
        records["FR2"].source_sha256,
    ]
    assert [case.tllm_text_sha256 for case in export.cases] == [
        records["EN1"].translation_sha256,
        records["ES3"].translation_sha256,
        records["FR2"].translation_sha256,
    ]

    workbook = load_workbook(output)
    try:
        review = workbook["Review"]
        assert [review.cell(row, 1).value for row in range(2, 5)] == [
            "EN1",
            "ES3",
            "FR2",
        ]
        assert [review.cell(row, 3).value for row in range(2, 5)] == [
            store.read_bytes(records[case_id].source_sha256).decode("utf-8")
            for case_id in ("EN1", "ES3", "FR2")
        ]
        assert [review.cell(row, 4).value for row in range(2, 5)] == [
            store.read_bytes(records[case_id].translation_sha256).decode("utf-8")
            for case_id in ("EN1", "ES3", "FR2")
        ]
    finally:
        workbook.close()


def test_export_includes_nmt_only_when_its_manifest_is_requested(
    tmp_path: Path,
) -> None:
    store = ArtifactStore(tmp_path / "objects")
    tllm, nmt = _manifests(store)
    output = tmp_path / "review.xlsx"

    export_sha256 = export_translation_review(
        store=store,
        tllm_manifest=tllm,
        nmt_manifest=nmt,
        destination=output,
        review_policy_id="medical-review-v1",
    )

    export = TranslationReviewExport.model_validate_json(
        store.read_bytes(export_sha256), strict=True
    )
    nmt_records = {record.source_case_id: record for record in nmt.records}
    assert export.nmt_recipe_sha256 == nmt.recipe_sha256
    assert [case.nmt_text_sha256 for case in export.cases] == [
        nmt_records["EN1"].translation_sha256,
        nmt_records["ES3"].translation_sha256,
        nmt_records["FR2"].translation_sha256,
    ]

    workbook = load_workbook(output)
    try:
        review = workbook["Review"]
        assert [review.cell(row, 11).value for row in range(2, 5)] == [
            store.read_bytes(nmt_records[case_id].translation_sha256).decode("utf-8")
            for case_id in ("EN1", "ES3", "FR2")
        ]
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("change", "message"),
    [
        (
            lambda store, nmt: nmt.model_copy(update={"selection_id": "selection-2"}),
            "selection ID",
        ),
        (
            lambda store, nmt: nmt.model_copy(update={"records": nmt.records[:-1]}),
            "case sets",
        ),
        (
            lambda store, nmt: nmt.model_copy(
                update={
                    "records": (
                        nmt.records[0].model_copy(
                            update={
                                "source_sha256": store.put_bytes(b"different source")
                            }
                        ),
                        *nmt.records[1:],
                    )
                }
            ),
            "source hashes",
        ),
        (
            lambda store, nmt: nmt.model_copy(
                update={
                    "records": (
                        nmt.records[0].model_copy(update={"source_language": "en"}),
                        *nmt.records[1:],
                    )
                }
            ),
            "source languages",
        ),
    ],
)
def test_export_rejects_incompatible_nmt_manifest(
    tmp_path: Path,
    change: object,
    message: str,
) -> None:
    store = ArtifactStore(tmp_path / "objects")
    tllm, nmt = _manifests(store)

    with pytest.raises(ValueError, match=message):
        export_translation_review(
            store=store,
            tllm_manifest=tllm,
            nmt_manifest=change(store, nmt),  # type: ignore[operator]
            destination=tmp_path / "review.xlsx",
            review_policy_id="medical-review-v1",
        )


def test_export_decodes_artifact_text_as_strict_utf8(tmp_path: Path) -> None:
    store = ArtifactStore(tmp_path / "objects")
    tllm, _ = _manifests(store)
    invalid = tllm.records[0].model_copy(
        update={"translation_sha256": store.put_bytes(b"\xff")}
    )

    with pytest.raises(UnicodeDecodeError):
        export_translation_review(
            store=store,
            tllm_manifest=tllm.model_copy(
                update={"records": (invalid, *tllm.records[1:])}
            ),
            destination=tmp_path / "review.xlsx",
            review_policy_id="medical-review-v1",
        )
