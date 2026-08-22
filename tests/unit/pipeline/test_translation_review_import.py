from collections.abc import Callable
from datetime import UTC, datetime
from decimal import Decimal
from itertools import product
from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from pydantic import ValidationError

from phentrieve_benchmark.artifacts.store import ArtifactStore
from phentrieve_benchmark.models.review import ReviewRecord
from phentrieve_benchmark.models.translation import (
    TranslationManifest,
    TranslationRecord,
    TranslationStatus,
)
from phentrieve_benchmark.models.translation_review import (
    TranslationReviewDiff,
    TranslationReviewImportManifest,
    TranslationReviewRecord,
)
from phentrieve_benchmark.pipeline.translation_review import (
    WorkbookValidationError,
    export_translation_review,
    import_translation_review,
)


def _translation_manifest(
    store: ArtifactStore,
    *,
    model: str = "general/translation-llm",
    noncanonical_field: str | None = None,
) -> TranslationManifest:
    cases = (
        (
            "FR1",
            "fr",
            "Rapport français.",
            "Deutscher TLLM-Text.",
            "Deutscher NMT-Text.",
        ),
        (
            "EN1",
            "en",
            "English report.",
            "Unveränderter Text.",
            "Unveränderter NMT-Text.",
        ),
    )

    def noncanonical_texts(
        case_id: str, source: str, tllm: str, nmt: str
    ) -> tuple[str, str, str]:
        if case_id != "EN1":
            return source, tllm, nmt
        if noncanonical_field == "source_text":
            source = "Cafe\u0301 source."
        elif noncanonical_field == "tllm_text":
            tllm = "U\u0308bersetzung."
        elif noncanonical_field == "nmt_text":
            nmt = "U\u0308bersetzung."
        return source, tllm, nmt

    normalized_cases = (
        (case_id, language, *noncanonical_texts(case_id, source, tllm, nmt))
        for case_id, language, source, tllm, nmt in cases
    )
    records = tuple(
        TranslationRecord(
            translation_id=f"{model}-{case_id}",
            selection_id="selection-1",
            source_case_id=case_id,
            source_language=language,  # type: ignore[arg-type]
            target_language="de",
            source_sha256=store.put_bytes(source.encode()),
            translation_sha256=store.put_bytes(
                (tllm if model == "general/translation-llm" else nmt).encode()
            ),
            provider="google-cloud-translation",
            api_version="v3",
            model=model,  # type: ignore[arg-type]
            project_id="phentrieve",
            location="global",
            created_at=datetime(2026, 8, 22, tzinfo=UTC),
            input_codepoints=len(source),
            output_codepoints=len(tllm if model == "general/translation-llm" else nmt),
            price_per_million_input_characters=Decimal("0"),
            estimated_max_cost=Decimal("0"),
            status=TranslationStatus.TRANSLATED,
            checks=(),
        )
        for case_id, language, source, tllm, nmt in normalized_cases
    )
    manifest = TranslationManifest(
        selection_id="selection-1",
        selection_sha256="a" * 64,
        recipe_sha256=("b" if model == "general/translation-llm" else "c") * 64,
        records=records,
    )
    store.put_bytes(manifest.canonical_bytes())
    return manifest


def _completed_workbook(
    tmp_path: Path,
    *,
    include_nmt: bool = False,
    noncanonical_field: str | None = None,
) -> tuple[ArtifactStore, Path]:
    store = ArtifactStore(tmp_path / "objects")
    workbook_path = tmp_path / "review.xlsx"
    nmt_manifest = (
        _translation_manifest(
            store,
            model="general/nmt",
            noncanonical_field=noncanonical_field,
        )
        if include_nmt
        else None
    )
    export_translation_review(
        store=store,
        tllm_manifest=_translation_manifest(
            store, noncanonical_field=noncanonical_field
        ),
        nmt_manifest=nmt_manifest,
        destination=workbook_path,
        review_policy_id="medical-review-v1",
    )
    workbook = load_workbook(workbook_path)
    try:
        instructions = workbook["Anleitung"]
        instructions["B10"] = "reviewer-1"
        instructions["B11"] = "medical translator"
        instructions["B12"] = "English, French, German"
        instructions["B13"] = "2026-08-22"
        review = workbook["Review"]
        review["F2"] = "unverändert akzeptiert"
        review["G2"] = "keine"
        review["E3"] = "Korrigierter Text\n"
        review["F3"] = "korrigiert akzeptiert"
        review["G3"] = "vorhanden"
        review["H3"] = "Terminologie"
        review["I3"] = "Terminologie wurde korrigiert."
        workbook.save(workbook_path)
    finally:
        workbook.close()
    return store, workbook_path


class _RecordingStore(ArtifactStore):
    def __init__(self, root: Path, *, fail_at: int | None = None) -> None:
        super().__init__(root)
        self.values: list[bytes] = []
        self.fail_at = fail_at

    def put_bytes(self, value: bytes) -> str:
        self.values.append(value)
        if self.fail_at == len(self.values):
            raise OSError("injected object write failure")
        return super().put_bytes(value)


def _mutate_workbook(path: Path, change: Callable[[Any], None]) -> None:
    workbook = load_workbook(path)
    try:
        change(workbook)
        workbook.save(path)
    finally:
        workbook.close()


def _assert_rejected_without_writes(
    store: ArtifactStore,
    workbook_path: Path,
    *,
    fields: set[str],
) -> WorkbookValidationError:
    recording = _RecordingStore(store.root)
    with pytest.raises(WorkbookValidationError) as caught:
        import_translation_review(store=recording, workbook_path=workbook_path)
    assert recording.values == []
    assert fields <= {issue.field for issue in caught.value.issues}
    assert all(
        issue.sheet and issue.field and issue.message for issue in caught.value.issues
    )
    return caught.value


def test_import_publishes_canonical_review_manifest(tmp_path: Path) -> None:
    store, workbook_path = _completed_workbook(tmp_path)

    manifest_sha256 = import_translation_review(
        store=store,
        workbook_path=workbook_path,
    )
    manifest = TranslationReviewImportManifest.model_validate_json(
        store.read_bytes(manifest_sha256), strict=True
    )

    assert [entry.source_case_id for entry in manifest.entries] == ["EN1", "FR1"]
    assert store.read_bytes(manifest.entries[1].proposed_text_sha256) == (
        b"Korrigierter Text\n"
    )
    assert (
        import_translation_review(
            store=store,
            workbook_path=workbook_path,
        )
        == manifest_sha256
    )


@pytest.mark.parametrize(
    ("include_nmt", "coordinate", "field"),
    [
        (False, "B2", "source_language"),
        (False, "C2", "source_text"),
        (False, "D2", "tllm_text"),
        (True, "K2", "nmt_text"),
    ],
)
def test_import_rejects_tampered_immutable_cells_before_writing(
    tmp_path: Path,
    include_nmt: bool,
    coordinate: str,
    field: str,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path, include_nmt=include_nmt)
    _mutate_workbook(
        workbook_path,
        lambda workbook: setattr(workbook["Review"][coordinate], "value", "tampered"),
    )

    error = _assert_rejected_without_writes(store, workbook_path, fields={field})

    assert f"Review row 2 case EN1 field {field}" in str(error)


def test_import_requires_export_id_to_resolve_in_the_store(tmp_path: Path) -> None:
    store, workbook_path = _completed_workbook(tmp_path)
    _mutate_workbook(
        workbook_path,
        lambda workbook: setattr(workbook["Anleitung"]["B15"], "value", "0" * 64),
    )

    error = _assert_rejected_without_writes(
        store, workbook_path, fields={"export_sha256"}
    )

    assert "valid canonical export" in str(error)


@pytest.mark.parametrize(
    ("field", "include_nmt"),
    [
        ("source_text", False),
        ("tllm_text", False),
        ("nmt_text", True),
    ],
)
def test_import_rejects_noncanonical_authoritative_text_artifacts(
    tmp_path: Path, field: str, include_nmt: bool
) -> None:
    store, workbook_path = _completed_workbook(
        tmp_path,
        include_nmt=include_nmt,
        noncanonical_field=field,
    )

    error = _assert_rejected_without_writes(store, workbook_path, fields={field})

    assert "canonical text bytes" in str(error)


@pytest.mark.parametrize("case_change", ["missing", "extra", "duplicate"])
def test_import_requires_the_exact_export_case_set(
    tmp_path: Path, case_change: str
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)

    def change(workbook: Any) -> None:
        review = workbook["Review"]
        if case_change == "missing":
            review.delete_rows(3)
        elif case_change == "extra":
            review.append([review.cell(2, column).value for column in range(1, 11)])
            review["A4"] = "EXTRA"
        else:
            review["A3"] = "EN1"

    _mutate_workbook(workbook_path, change)

    _assert_rejected_without_writes(store, workbook_path, fields={"source_case_id"})


@pytest.mark.parametrize(
    ("sheet", "coordinate", "value", "field"),
    [
        ("Anleitung", "B16", "other-selection", "selection_id"),
        ("Anleitung", "B17", "other-policy", "review_policy_id"),
        ("Anleitung", "B10", "", "reviewer_id"),
        ("Anleitung", "B11", "", "reviewer_qualification"),
        ("Anleitung", "B12", "", "reviewed_languages"),
        ("Anleitung", "B13", "2026-02-30", "review_date"),
        ("Review", "E2", "", "proposed_text"),
        ("Review", "F2", "not-a-decision", "decision"),
        ("Review", "G2", "not-a-change", "clinical_change"),
        ("Review", "H3", "not-a-category", "clinical_change_category"),
    ],
)
def test_import_rejects_invalid_metadata_and_controlled_values(
    tmp_path: Path,
    sheet: str,
    coordinate: str,
    value: str,
    field: str,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)
    _mutate_workbook(
        workbook_path,
        lambda workbook: setattr(workbook[sheet][coordinate], "value", value),
    )

    error = _assert_rejected_without_writes(store, workbook_path, fields={field})

    assert "row" in str(error) and f"field {field}" in str(error)


_DECISION_COMBINATIONS = tuple(
    combination
    for combination in product(
        (
            "unverändert akzeptiert",
            "korrigiert akzeptiert",
            "Rückfrage",
            "abgelehnt",
        ),
        (False, True),
        ("keine", "vorhanden"),
        (False, True),
        (False, True),
    )
    if not (
        (
            combination[0] == "unverändert akzeptiert"
            and not combination[1]
            and combination[2] == "keine"
            and not combination[3]
            and not combination[4]
        )
        or (
            combination[0] == "korrigiert akzeptiert"
            and combination[1]
            and (
                (
                    combination[2] == "keine"
                    and not combination[3]
                    and not combination[4]
                )
                or (combination[2] == "vorhanden" and combination[3] and combination[4])
            )
        )
        or (
            combination[0] in {"Rückfrage", "abgelehnt"}
            and combination[2] == "vorhanden"
            and combination[3]
            and combination[4]
        )
    )
)


@pytest.mark.parametrize(
    ("decision", "changed", "clinical_change", "has_category", "has_rationale"),
    _DECISION_COMBINATIONS,
)
def test_import_rejects_every_forbidden_decision_combination(
    tmp_path: Path,
    decision: str,
    changed: bool,
    clinical_change: str,
    has_category: bool,
    has_rationale: bool,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)

    def change(workbook: Any) -> None:
        review = workbook["Review"]
        review["E2"] = "Geänderter Text." if changed else "Unveränderter Text."
        review["F2"] = decision
        review["G2"] = clinical_change
        review["H2"] = "Terminologie" if has_category else ""
        review["I2"] = "Klinische Begründung." if has_rationale else ""

    _mutate_workbook(workbook_path, change)

    _assert_rejected_without_writes(store, workbook_path, fields={"decision"})


@pytest.mark.parametrize("malformation", ["extra_sheet", "formula"])
def test_import_wraps_structural_parser_failures(
    tmp_path: Path, malformation: str
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)

    def change(workbook: Any) -> None:
        if malformation == "extra_sheet":
            workbook.create_sheet("Unexpected")
        else:
            workbook["Review"]["E2"] = "=1+1"

    _mutate_workbook(workbook_path, change)

    expected_field = "proposed_text" if malformation == "formula" else "workbook"
    error = _assert_rejected_without_writes(
        store, workbook_path, fields={expected_field}
    )
    assert malformation.split("_")[0] in str(error).casefold()
    if malformation == "formula":
        assert "Review row 2 case EN1 field proposed_text" in str(error)


def test_import_maps_non_string_metadata_errors_to_their_cell(tmp_path: Path) -> None:
    store, workbook_path = _completed_workbook(tmp_path)
    _mutate_workbook(
        workbook_path,
        lambda workbook: setattr(workbook["Anleitung"]["B13"], "value", 45_526),
    )

    error = _assert_rejected_without_writes(
        store, workbook_path, fields={"review_date"}
    )

    assert "Anleitung row 13 case - field review_date" in str(error)


def test_import_keeps_review_sheet_for_ambiguous_non_string_coordinate(
    tmp_path: Path,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)
    _mutate_workbook(
        workbook_path,
        lambda workbook: setattr(workbook["Review"]["B3"], "value", 123),
    )

    error = _assert_rejected_without_writes(
        store, workbook_path, fields={"source_language"}
    )

    assert "Review row 3 case FR1 field source_language" in str(error)


def test_import_aggregates_all_formulas_and_other_metadata_errors(
    tmp_path: Path,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)

    def change(workbook: Any) -> None:
        workbook["Anleitung"]["B10"] = ""
        workbook["Anleitung"]["B13"] = "bad-date"
        workbook["Review"]["E2"] = "=1+1"
        workbook["Review"]["J3"] = "=2+2"

    _mutate_workbook(workbook_path, change)

    error = _assert_rejected_without_writes(
        store,
        workbook_path,
        fields={"reviewer_id", "review_date", "proposed_text", "reviewer_comment"},
    )

    formula_issues = [issue for issue in error.issues if "formula" in issue.message]
    assert [(issue.row, issue.case_id, issue.field) for issue in formula_issues] == [
        (2, "EN1", "proposed_text"),
        (3, "FR1", "reviewer_comment"),
    ]


def test_import_aggregates_non_string_consumed_cells_and_semantic_errors(
    tmp_path: Path,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)

    def change(workbook: Any) -> None:
        workbook["Anleitung"]["B13"] = 45_526
        workbook["Review"]["B2"] = 1
        workbook["Review"]["B3"] = 2
        workbook["Review"]["F2"] = "bad-decision"

    _mutate_workbook(workbook_path, change)

    error = _assert_rejected_without_writes(
        store,
        workbook_path,
        fields={"review_date", "source_language", "decision"},
    )

    non_string_issues = [
        issue for issue in error.issues if issue.message == "must contain a string"
    ]
    assert [
        (issue.sheet, issue.row, issue.case_id, issue.field)
        for issue in non_string_issues
    ] == [
        ("Anleitung", 13, None, "review_date"),
        ("Review", 2, "EN1", "source_language"),
        ("Review", 3, "FR1", "source_language"),
    ]
    assert any(issue.field == "decision" for issue in error.issues)


def test_import_rejects_utf16_cell_length_overflow(tmp_path: Path) -> None:
    store, workbook_path = _completed_workbook(tmp_path)
    _mutate_workbook(
        workbook_path,
        lambda workbook: setattr(workbook["Review"]["E2"], "value", "😀" * 16_384),
    )

    error = _assert_rejected_without_writes(
        store, workbook_path, fields={"proposed_text"}
    )

    assert "32,767 UTF-16" in str(error)


def test_import_reports_all_detectable_validation_failures(tmp_path: Path) -> None:
    store, workbook_path = _completed_workbook(tmp_path)

    def change(workbook: Any) -> None:
        workbook["Anleitung"]["B10"] = ""
        workbook["Anleitung"]["B13"] = "bad-date"
        workbook["Review"]["C2"] = "tampered source"
        workbook["Review"]["F2"] = "bad-decision"

    _mutate_workbook(workbook_path, change)

    error = _assert_rejected_without_writes(
        store,
        workbook_path,
        fields={"reviewer_id", "review_date", "source_text", "decision"},
    )

    assert len(error.issues) >= 4


def test_import_stores_the_reachable_review_graph_with_manifest_last(
    tmp_path: Path,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)
    recording = _RecordingStore(store.root)

    manifest_sha256 = import_translation_review(
        store=recording, workbook_path=workbook_path
    )
    manifest = TranslationReviewImportManifest.model_validate_json(
        recording.values[-1], strict=True
    )

    assert manifest.sha256() == manifest_sha256
    assert len(recording.values) == 9
    for entry in manifest.entries:
        record = TranslationReviewRecord.model_validate_json(
            store.read_bytes(entry.record_sha256), strict=True
        )
        projection = ReviewRecord.model_validate_json(
            store.read_bytes(entry.review_record_sha256), strict=True
        )
        diff = TranslationReviewDiff.model_validate_json(
            store.read_bytes(entry.diff_sha256), strict=True
        )
        assert projection.review_id == f"translation-review:{entry.record_sha256}"
        assert projection.subject_sha256 == entry.proposed_text_sha256
        assert record.proposed_text_sha256 == entry.proposed_text_sha256
        assert diff.proposed_text_sha256 == entry.proposed_text_sha256
    assert not any(path.name == "latest" for path in store.root.rglob("*"))


def test_import_does_not_publish_a_manifest_after_an_object_write_failure(
    tmp_path: Path,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)
    failing = _RecordingStore(store.root, fail_at=3)

    with pytest.raises(OSError, match="injected"):
        import_translation_review(store=failing, workbook_path=workbook_path)

    for value in failing.values:
        with pytest.raises(ValidationError):
            TranslationReviewImportManifest.model_validate_json(value, strict=True)

    manifest_sha256 = import_translation_review(
        store=store, workbook_path=workbook_path
    )
    TranslationReviewImportManifest.model_validate_json(
        store.read_bytes(manifest_sha256), strict=True
    )


@pytest.mark.parametrize("profile_change", ["add_nmt", "remove_nmt"])
def test_import_requires_nmt_column_presence_to_match_the_export(
    tmp_path: Path, profile_change: str
) -> None:
    include_nmt = profile_change == "remove_nmt"
    store, workbook_path = _completed_workbook(tmp_path, include_nmt=include_nmt)

    def change(workbook: Any) -> None:
        review = workbook["Review"]
        if profile_change == "add_nmt":
            review["K1"] = "NMT-Vergleich"
            review["K2"] = "unexpected NMT"
            review["K3"] = "unexpected NMT"
        else:
            review.delete_cols(11)

    _mutate_workbook(workbook_path, change)

    _assert_rejected_without_writes(store, workbook_path, fields={"nmt_text"})


def test_import_identity_ignores_formatting_but_includes_semantic_revision(
    tmp_path: Path,
) -> None:
    store, workbook_path = _completed_workbook(tmp_path)
    first_sha256 = import_translation_review(store=store, workbook_path=workbook_path)
    first_bytes = store.read_bytes(first_sha256)

    def format_only(workbook: Any) -> None:
        review = workbook["Review"]
        review.column_dimensions["E"].width = 72
        review.row_dimensions[2].height = 120
        review["L2"].font = Font(italic=True)
        review["A1000"].alignment = Alignment(wrap_text=False)

    _mutate_workbook(workbook_path, format_only)
    assert (
        import_translation_review(store=store, workbook_path=workbook_path)
        == first_sha256
    )

    _mutate_workbook(
        workbook_path,
        lambda workbook: setattr(
            workbook["Review"]["J2"], "value", "Second semantic review."
        ),
    )
    second_sha256 = import_translation_review(store=store, workbook_path=workbook_path)

    assert second_sha256 != first_sha256
    assert store.read_bytes(first_sha256) == first_bytes
