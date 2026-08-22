from pathlib import Path

import pytest
from openpyxl import load_workbook

from phentrieve_benchmark.models.translation_review import (
    TranslationReviewExport,
    TranslationReviewExportCase,
)
from phentrieve_benchmark.review.translation_workbook import (
    CATEGORY_VALUES,
    CLINICAL_CHANGE_VALUES,
    DECISION_VALUES,
    REVIEW_HEADERS,
    WorkbookCase,
    read_review_workbook,
    write_review_workbook,
)


def _export(*, include_nmt: bool = False) -> TranslationReviewExport:
    nmt_sha256 = "3" * 64 if include_nmt else None
    return TranslationReviewExport(
        selection_id="e3c-selection-v1",
        review_policy_id="medical-review-v1",
        nmt_recipe_sha256="4" * 64 if include_nmt else None,
        cases=(
            TranslationReviewExportCase(
                source_case_id="case-1",
                source_language="en",
                source_text_sha256="1" * 64,
                tllm_text_sha256="2" * 64,
                nmt_text_sha256=nmt_sha256,
            ),
            TranslationReviewExportCase(
                source_case_id="case-2",
                source_language="fr",
                source_text_sha256="5" * 64,
                tllm_text_sha256="6" * 64,
                nmt_text_sha256="7" * 64 if include_nmt else None,
            ),
        ),
    )


def _cases(*, include_nmt: bool = False) -> tuple[WorkbookCase, ...]:
    return (
        WorkbookCase(
            source_case_id="case-1",
            source_language="en",
            source_text="Original eins",
            tllm_text="TLLM eins",
            nmt_text="NMT eins" if include_nmt else None,
        ),
        WorkbookCase(
            source_case_id="case-2",
            source_language="fr",
            source_text="Original zwei",
            tllm_text="TLLM zwei",
            nmt_text="NMT zwei" if include_nmt else None,
        ),
    )


def _write(path: Path, *, include_nmt: bool = False) -> None:
    write_review_workbook(
        path, _export(include_nmt=include_nmt), _cases(include_nmt=include_nmt)
    )


def test_writer_creates_the_minimal_review_profile(tmp_path: Path) -> None:
    output = tmp_path / "review.xlsx"

    _write(output)

    workbook = load_workbook(output)
    try:
        assert workbook.sheetnames == ["Anleitung", "Review"]
        review = workbook["Review"]
        assert review.freeze_panes == "A2"
        assert [cell.value for cell in review[1]] == list(REVIEW_HEADERS)
        assert review["E2"].value == "TLLM eins"
        assert review["A2"].protection.locked is True
        assert review["E2"].protection.locked is False
        assert review.protection.sheet is True
        assert review.protection.selectLockedCells is True
        assert review.protection.autoFilter is False
        assert review.auto_filter.ref == "A1:J3"
        assert review["A2"].alignment.wrap_text is True
        assert review["A2"].alignment.vertical == "top"

        validations = {
            validation.formula1 for validation in review.data_validations.dataValidation
        }
        assert validations == {
            '"' + ",".join(DECISION_VALUES) + '"',
            '"' + ",".join(CLINICAL_CHANGE_VALUES) + '"',
            '"' + ",".join(CATEGORY_VALUES) + '"',
        }

        instructions = workbook["Anleitung"]
        assert instructions["B3"].value == _export().sha256()
        assert instructions["B10"].number_format == "@"
        assert instructions["B10"].protection.locked is False
        assert any(
            validation.type == "custom" and "ISTEXT(B10)" in validation.formula1
            for validation in instructions.data_validations.dataValidation
        )
        instruction_text = "\n".join(
            str(instructions[f"A{row}"].value) for row in range(12, 18)
        )
        assert "Metadaten" in instruction_text
        assert "Originaltext" in instruction_text
        assert all(decision in instruction_text for decision in DECISION_VALUES)
        assert "vorhanden" in instruction_text
        assert "Hauptkategorie" in instruction_text
        assert "Änderungsbegründung" in instruction_text
        assert "unverändert akzeptiert" in instruction_text
        assert "Rückfrage" in instruction_text
    finally:
        workbook.close()


def test_writer_adds_nmt_only_for_nmt_exports(tmp_path: Path) -> None:
    output = tmp_path / "review-nmt.xlsx"

    _write(output, include_nmt=True)

    workbook = load_workbook(output)
    try:
        review = workbook["Review"]
        assert [cell.value for cell in review[1]] == [*REVIEW_HEADERS, "NMT-Vergleich"]
        assert review["K2"].value == "NMT eins"
        assert review["K2"].protection.locked is True
    finally:
        workbook.close()


def test_writer_rejects_values_over_excel_utf16_cell_limit(tmp_path: Path) -> None:
    output = tmp_path / "review.xlsx"
    oversized_cases = (
        WorkbookCase(
            source_case_id="case-1",
            source_language="en",
            source_text="😀" * 16_384,
            tllm_text="TLLM eins",
        ),
        _cases()[1],
    )

    with pytest.raises(ValueError, match="32,767 UTF-16"):
        write_review_workbook(output, _export(), oversized_cases)

    assert not output.exists()


def test_reader_returns_string_metadata_and_row_values(tmp_path: Path) -> None:
    output = tmp_path / "review.xlsx"
    _write(output)
    workbook = load_workbook(output)
    try:
        instructions = workbook["Anleitung"]
        instructions["B7"] = "reviewer-1"
        instructions["B8"] = "medical translator"
        instructions["B9"] = "English, French, German"
        instructions["B10"] = "2026-08-22"
        review = workbook["Review"]
        review["F2"] = "unverändert akzeptiert"
        review["G2"] = "keine"
        review["H2"] = ""
        review["I2"] = ""
        review["J2"] = ""
        workbook.save(output)
    finally:
        workbook.close()

    parsed = read_review_workbook(output)

    assert parsed.export_sha256 == _export().sha256()
    assert parsed.reviewer_id == "reviewer-1"
    assert parsed.review_date == "2026-08-22"
    assert parsed.rows[0].source_case_id == "case-1"
    assert parsed.rows[0].proposed_text == "TLLM eins"
    assert parsed.rows[0].decision == "unverändert akzeptiert"
    assert parsed.rows[0].clinical_change_category == ""


@pytest.mark.parametrize("wrong_suffix", [".xls"])
def test_reader_requires_xlsx_suffix(tmp_path: Path, wrong_suffix: str) -> None:
    source = tmp_path / f"review{wrong_suffix}"
    source.write_bytes(b"not a workbook")

    with pytest.raises(ValueError, match=r"\.xlsx"):
        read_review_workbook(source)


def test_reader_rejects_unexpected_profile_content(tmp_path: Path) -> None:
    output = tmp_path / "review.xlsx"
    _write(output)
    workbook = load_workbook(output)
    try:
        workbook.create_sheet("Unexpected")
        workbook.save(output)
    finally:
        workbook.close()

    with pytest.raises(ValueError, match="exactly"):
        read_review_workbook(output)


def test_reader_rejects_formula_and_non_string_values(tmp_path: Path) -> None:
    output = tmp_path / "review.xlsx"
    _write(output)
    workbook = load_workbook(output)
    try:
        workbook["Review"]["E2"] = "=1+1"
        workbook.save(output)
    finally:
        workbook.close()

    with pytest.raises(ValueError, match="formula"):
        read_review_workbook(output)

    workbook = load_workbook(output)
    try:
        workbook["Review"]["E2"] = 2
        workbook.save(output)
    finally:
        workbook.close()

    with pytest.raises(ValueError, match="string"):
        read_review_workbook(output)
