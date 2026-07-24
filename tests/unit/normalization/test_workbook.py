import io
import zipfile

import pytest
from openpyxl import Workbook

from phentrieve_benchmark.acquisition.recipes import WorkbookLimits
from phentrieve_benchmark.normalization.workbook import (
    WorkbookValidationError,
    open_validated_workbook,
)


def _workbook(*, formula: bool = False) -> bytes:
    output = io.BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GSC Manual Annotations "
    sheet.append(["value"])
    sheet.append(["=1+1" if formula else "plain"])
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _limits(**updates: int) -> WorkbookLimits:
    values = {
        "maximum_member_count": 100,
        "maximum_member_bytes": 1_000_000,
        "maximum_expanded_bytes": 5_000_000,
        "maximum_compression_ratio": 100,
    }
    values.update(updates)
    return WorkbookLimits(**values)


def _add_member(body: bytes, name: str, value: bytes = b"x") -> bytes:
    source = zipfile.ZipFile(io.BytesIO(body))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w") as target:
        for info in source.infolist():
            target.writestr(info, source.read(info))
        target.writestr(name, value)
    return output.getvalue()


def test_opens_minimal_workbook_and_preserves_exact_sheet_name() -> None:
    workbook = open_validated_workbook(_workbook(), limits=_limits())
    try:
        assert workbook.sheetnames == ["GSC Manual Annotations "]
        assert "GSC Manual Annotations" not in workbook.sheetnames
    finally:
        workbook.close()


@pytest.mark.parametrize(
    ("limits", "message"),
    [
        (_limits(maximum_member_count=1), "member count"),
        (_limits(maximum_member_bytes=1), "member size"),
        (_limits(maximum_expanded_bytes=1), "expanded"),
        (_limits(maximum_compression_ratio=1), "compression ratio"),
    ],
)
def test_enforces_nested_archive_limits(
    limits: WorkbookLimits, message: str
) -> None:
    with pytest.raises(WorkbookValidationError, match=message):
        open_validated_workbook(_workbook(), limits=limits)


@pytest.mark.parametrize(
    "name",
        [
            "../escape.xml",
            "xl/vbaProject.bin",
        "xl/activeX/control.bin",
        "xl/embeddings/object.bin",
        "xl/externalLinks/externalLink1.xml",
        "xl/connections.xml",
        "xl/queryTables/query1.xml",
    ],
)
def test_rejects_unsafe_or_active_package_members(name: str) -> None:
    with pytest.raises(WorkbookValidationError):
        open_validated_workbook(_add_member(_workbook(), name), limits=_limits())


def test_rejects_formulas() -> None:
    with pytest.raises(WorkbookValidationError, match="formula"):
        open_validated_workbook(_workbook(formula=True), limits=_limits())
