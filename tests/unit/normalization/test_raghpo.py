import io
from pathlib import Path

from openpyxl import Workbook

from phentrieve_benchmark.acquisition.recipes import (
    ExpectedCount,
    ExpectedTable,
    load_source_recipe,
    load_target_recipe,
)
from phentrieve_benchmark.normalization.raghpo import normalize_raghpo_target

ROOT = Path(__file__).parents[3]


def _workbook(*, csc_note: str = "Note A") -> bytes:
    output = io.BytesIO()
    workbook = Workbook()
    csc_input = workbook.active
    csc_input.title = "CSC Input"
    csc_input.append(["Case", "clinical_note"])
    csc_input.append([1, csc_note])
    csc_manual = workbook.create_sheet("CSC Manual Annotations")
    csc_manual.append(["Patient ID", "hpo_description", "hpo_term"])
    csc_manual.append([1, "first", " HP:0000001,HP:0000002 "])
    gsc_input = workbook.create_sheet("GSC Input")
    gsc_input.append(["patient_id", "ID", "clinical_note"])
    gsc_input.append(["patient:1", "A/B", "Note G"])
    gsc_manual = workbook.create_sheet("GSC Manual Annotations ")
    gsc_manual.append(
        ["Patient ID", "ID", "hpo_description", "hpo_term", "Category"]
    )
    gsc_manual.append(["patient:1", "A/B", "finding", "HP:0000001", "cat"])
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _target(target: str):
    loaded = load_target_recipe(
        ROOT / f"datasets/raghpo/{target}/dataset.yaml"
    ).value
    if target == "csc":
        tables = (
            ExpectedTable(
                source_path="RAG-HPO Tests and Data Analysis copy.xlsx",
                sheet_name="CSC Input",
                columns=("Case", "clinical_note"),
                data_rows=1,
            ),
            ExpectedTable(
                source_path="RAG-HPO Tests and Data Analysis copy.xlsx",
                sheet_name="CSC Manual Annotations",
                columns=("Patient ID", "hpo_description", "hpo_term"),
                data_rows=1,
            ),
        )
        annotations = 2
    else:
        tables = (
            ExpectedTable(
                source_path="RAG-HPO Tests and Data Analysis copy.xlsx",
                sheet_name="GSC Input",
                columns=("patient_id", "ID", "clinical_note"),
                data_rows=1,
            ),
            ExpectedTable(
                source_path="RAG-HPO Tests and Data Analysis copy.xlsx",
                sheet_name="GSC Manual Annotations ",
                columns=(
                    "Patient ID",
                    "ID",
                    "hpo_description",
                    "hpo_term",
                    "Category",
                ),
                data_rows=1,
            ),
        )
        annotations = 1
    return loaded.model_copy(
        update={
            "expected_tables": tables,
            "expected_counts": (
                ExpectedCount(name="documents", count=1),
                ExpectedCount(name="annotations", count=annotations),
            ),
        }
    )


def test_csc_splits_ascii_comma_and_preserves_empty_evidence() -> None:
    source = load_source_recipe(ROOT / "datasets/raghpo/source.yaml").value
    result = normalize_raghpo_target(
        workbook_bytes=_workbook(),
        source_recipe=source,
        target_recipe=_target("csc"),
    )
    assert len(result.documents) == 1
    assert len(result.annotation_sets[0].annotations) == 2
    assert all(
        not annotation.evidence_spans
        for annotation in result.annotation_sets[0].annotations
    )
    assert result.source_sidecar[0].raw_hpo_term == (
        " HP:0000001,HP:0000002 "
    )


def test_csc_uses_workbook_text_and_gsc_is_independent() -> None:
    source = load_source_recipe(ROOT / "datasets/raghpo/source.yaml").value
    csc = normalize_raghpo_target(
        workbook_bytes=_workbook(),
        source_recipe=source,
        target_recipe=_target("csc"),
    )
    assert csc.documents[0].text == "Note A"

    gsc = normalize_raghpo_target(
        workbook_bytes=_workbook(),
        source_recipe=source,
        target_recipe=_target("gsc"),
    )
    assert gsc.documents[0].source_case_id == "patient:1"
    assert ":patient%3A1:A%2FB:" in gsc.documents[0].document_id
    assert len(gsc.annotation_sets[0].annotations) == 1
