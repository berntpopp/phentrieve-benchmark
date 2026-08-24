from datetime import UTC, datetime
from decimal import Decimal
from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import load_workbook

from phentrieve_benchmark.artifacts.store import ArtifactStore
from phentrieve_benchmark.models.translation import (
    TranslationManifest,
    TranslationRecord,
    TranslationStatus,
)
from phentrieve_benchmark.pipeline.translation_review import (
    export_translation_review,
)
from phentrieve_benchmark.translation.pricing import load_translation_recipe

ROOT = Path(__file__).parents[2]
SNAPSHOT = (
    ROOT / "datasets/e3c-de/review/e3c-de-feasibility-30-v1"
)
RECIPE = ROOT / "datasets/e3c-de/translation-llm.yaml"


def _snapshot_bytes() -> dict[Path, bytes]:
    return {
        path.relative_to(SNAPSHOT): path.read_bytes()
        for path in SNAPSHOT.rglob("*.txt")
    }


def _snapshot_manifest(store: ArtifactStore) -> TranslationManifest:
    recipe = load_translation_recipe(RECIPE)
    records = []
    for case_directory in sorted(path for path in SNAPSHOT.iterdir() if path.is_dir()):
        source_path = next(case_directory.glob("source.*.txt"))
        source_language = source_path.name.split(".")[1]
        source = source_path.read_bytes()
        translation = (case_directory / "tllm.de.txt").read_bytes()
        records.append(
            TranslationRecord(
                translation_id=f"tracked-tllm-{case_directory.name}",
                selection_id="e3c-de-feasibility-30-v1",
                source_case_id=case_directory.name,
                source_language=source_language,  # type: ignore[arg-type]
                target_language="de",
                source_sha256=store.put_bytes(source),
                translation_sha256=store.put_bytes(translation),
                provider="google-cloud-translation",
                api_version="v3",
                model="general/translation-llm",
                project_id="phentrieve",
                location="us-central1",
                created_at=datetime(2026, 8, 22, tzinfo=UTC),
                input_codepoints=len(source.decode("utf-8")),
                output_codepoints=len(translation.decode("utf-8")),
                price_per_million_input_characters=Decimal("10"),
                estimated_max_cost=Decimal("0"),
                status=TranslationStatus.TRANSLATED,
                checks=(),
            )
        )
    return TranslationManifest(
        selection_id="e3c-de-feasibility-30-v1",
        selection_sha256=sha256(
            (
                ROOT
                / "datasets/e3c-de/selections/e3c-de-feasibility-30-v1.json"
            ).read_bytes()
        ).hexdigest(),
        recipe_sha256=recipe.sha256,
        records=tuple(records),
    )


def test_tracked_snapshot_exports_to_default_review_workbook(
    tmp_path: Path,
) -> None:
    before = _snapshot_bytes()
    store = ArtifactStore(tmp_path / "objects")
    manifest = _snapshot_manifest(store)
    destination = tmp_path / "review.xlsx"

    export_translation_review(
        store=store,
        tllm_manifest=manifest,
        destination=destination,
        review_policy_id="e3c:translation-review/v1",
    )

    workbook = load_workbook(destination, data_only=False)
    assert workbook.sheetnames == ["Anleitung", "Review"]
    review = workbook["Review"]
    assert review.max_row == 31
    headers = [cell.value for cell in review[1]]
    assert "NMT-Vergleich" not in headers
    source_by_case = {
        review.cell(row=row, column=1).value: review.cell(row=row, column=3).value
        for row in range(2, 32)
    }
    expected_french = before[Path("FR100185/source.fr.txt")].decode("utf-8")
    assert "é" in expected_french
    assert source_by_case["FR100185"] == expected_french
    assert _snapshot_bytes() == before


def test_language_filter_exports_only_that_language(tmp_path: Path) -> None:
    store = ArtifactStore(tmp_path / "objects")
    manifest = _snapshot_manifest(store)
    destination = tmp_path / "review-fr.xlsx"

    export_translation_review(
        store=store,
        tllm_manifest=manifest,
        destination=destination,
        review_policy_id="e3c:translation-review/v1",
        source_language="fr",
    )

    review = load_workbook(destination, data_only=False)["Review"]
    assert review.max_row == 11
    languages = {review.cell(row=row, column=2).value for row in range(2, 12)}
    assert languages == {"fr"}


def test_language_without_cases_is_rejected(tmp_path: Path) -> None:
    store = ArtifactStore(tmp_path / "objects")
    manifest = _snapshot_manifest(store)
    destination = tmp_path / "review-it.xlsx"

    with pytest.raises(ValueError, match="source language 'it'"):
        export_translation_review(
            store=store,
            tllm_manifest=manifest,
            destination=destination,
            review_policy_id="e3c:translation-review/v1",
            source_language="it",
        )

    assert not destination.exists()
